"""Aplicación de control de gastos con Tkinter.

Soporte para:
- Registro de facturas
- Registro de mano de obra (modo alternativo en el mismo formulario)
- Selección de proyecto al iniciar
- Exportación con totales y presupuestos separados

Modificaciones solicitadas:
- Dos casillas de presupuesto: Facturas y Mano de Obra
- Dos botones de exportación: Exportar Facturas / Exportar Mano de Obra
- Exportación filtrada por tipo
"""

from __future__ import annotations

import datetime as dt
import os
import tkinter as tk
from dataclasses import dataclass, asdict
from tkinter import filedialog, messagebox, ttk
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ARCHIVO_DATOS = "gastos_guardados.xlsx"
IVA = 0.16

COLUMNAS = [
    "Item",
    "Tipo",
    "Factura",
    "Descripción",
    "Unidad",
    "Cantidad",
    "Precio sin IVA (Bs)",
    "Precio con IVA (Bs)",
    "Precio sin IVA ($)",
    "Precio con IVA ($)",
    "Fecha",
    "Total sin IVA (Bs)",
    "Proveedor",
    "Semana",
    "Monto Manoobra ($)",
    "Gastos extras ($)",
]


# =========================
#   DATACLASS DEL REGISTRO
# =========================

@dataclass
class RegistroGasto:
    item: int
    tipo: str  # "Factura" o "Mano de obra"
    factura: str
    descripcion: str
    unidad: str
    cantidad: float
    precio_sin_iva_bs: float
    precio_con_iva_bs: float
    precio_sin_iva_usd: float
    precio_con_iva_usd: float
    fecha: str
    total_sin_iva: float
    proveedor: str
    proyecto: str
    semana: str = ""
    monto_manoobra_usd: float = 0.0
    gastos_extras_usd: float = 0.0

    @classmethod
    def from_dict(cls, data: dict[str, object]) -> "RegistroGasto":
        precio_sin_bs = float(data.get("Precio sin IVA (Bs)") or data.get("Precio (Bs)") or 0)
        precio_con_bs = float(data.get("Precio con IVA (Bs)") or round(precio_sin_bs * (1 + IVA), 2))
        precio_sin_usd = float(data.get("Precio sin IVA ($)") or data.get("Precio ($)") or 0)
        precio_con_usd = float(data.get("Precio con IVA ($)") or precio_sin_usd)
        cantidad = float(data.get("Cantidad", 0) or 0)
        total_sin = float(data.get("Total sin IVA (Bs)") or (precio_sin_bs * cantidad))

        return cls(
            item=int(data.get("Item", 0)),
            tipo=str(data.get("Tipo", "Factura")),
            factura=str(data.get("Factura", "")),
            descripcion=str(data.get("Descripción", "")),
            unidad=str(data.get("Unidad", "")),
            cantidad=cantidad,
            precio_sin_iva_bs=precio_sin_bs,
            precio_con_iva_bs=precio_con_bs,
            precio_sin_iva_usd=precio_sin_usd,
            precio_con_iva_usd=precio_con_usd,
            fecha=str(data.get("Fecha", "")),
            total_sin_iva=total_sin,
            proveedor=str(data.get("Proveedor", "")),
            proyecto=str(data.get("Proyecto", "")),
            semana=str(data.get("Semana", "")),
            monto_manoobra_usd=float(data.get("Monto Manoobra ($)", 0)),
            gastos_extras_usd=float(data.get("Gastos extras ($)", 0)),
        )

    def to_row(self) -> List[object]:
        return [
            self.item,
            self.tipo,
            self.factura,
            self.descripcion,
            self.unidad,
            self.cantidad,
            self.precio_sin_iva_bs,
            self.precio_con_iva_bs,
            self.precio_sin_iva_usd,
            self.precio_con_iva_usd,
            self.fecha,
            self.total_sin_iva,
            self.proveedor,
            self.semana,
            self.monto_manoobra_usd,
            self.gastos_extras_usd,
        ]


# =========================
#     CLASE PRINCIPAL
# =========================

class GestorGastos:
    def __init__(self) -> None:
        self.registros: List[RegistroGasto] = []
        self.precio_dolar: float = 0.0
        self.registro_en_edicion: Optional[int] = None
        self.proyecto: str = ""

        self.root = tk.Tk()
        self.root.title("Control de Gastos - Terra Caliza")
        self.root.geometry("1280x740")
        self.root.resizable(True, True)

        try:
            self.root.state("zoomed")
        except:
            pass

        self._configurar_estilos()
        self._crear_contenedor_scrollable()

        self._df_global = self._leer_archivo_datos()

        # Selección de proyecto antes de cargar la UI
        self._seleccionar_proyecto_modal()

        # Creación de la interfaz
        self._crear_componentes()
        self._cargar_datos_proyecto()
        self._refrescar_tabla()

    # ===== estilos =====
    def _configurar_estilos(self) -> None:
        style = ttk.Style(self.root)
        style.theme_use("clam")

        style.configure("Primary.TButton", background="#1976D2", foreground="white")
        style.map("Primary.TButton", background=[("active", "#1565C0")])

        style.configure("Success.TButton", background="#2E7D32", foreground="white")
        style.map("Success.TButton", background=[("active", "#1B5E20")])

        style.configure("Danger.TButton", background="#C62828", foreground="white")
        style.map("Danger.TButton", background=[("active", "#B71C1C")])

        style.configure("Header.TLabel", font=("Arial", 12, "bold"))

    # ===== contenedor scroll =====
    def _crear_contenedor_scrollable(self) -> None:
        self.canvas = tk.Canvas(self.root, borderwidth=0)
        self.scrollbar_y = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.scrollbar_x = ttk.Scrollbar(self.root, orient="horizontal", command=self.canvas.xview)

        self.canvas.configure(yscrollcommand=self.scrollbar_y.set, xscrollcommand=self.scrollbar_x.set)

        self.scrollbar_y.pack(side="right", fill="y")
        self.scrollbar_x.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.content_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        def _actualiza(_event=None):
            bbox = self.canvas.bbox("all")
            if bbox:
                self.canvas.configure(scrollregion=bbox)

        self.content_frame.bind("<Configure>", _actualiza)
        # =========================
    #    CREACIÓN DE UI
    # =========================

    def _crear_componentes(self) -> None:
        self._crear_frame_dolar()
        self._crear_formulario()
        self._crear_controles_tabla()
        self._crear_tabla()
        self._crear_resumen()
        self._crear_exportacion()
        self._crear_menu()

    def _crear_menu(self) -> None:
        menubar = tk.Menu(self.root)
        proyecto_menu = tk.Menu(menubar, tearoff=0)
        proyecto_menu.add_command(label="Cambiar proyecto", command=self._seleccionar_proyecto_modal)
        menubar.add_cascade(label="Proyecto", menu=proyecto_menu)
        self.root.config(menu=menubar)

    def _crear_frame_dolar(self) -> None:
        frame = ttk.LabelFrame(self.content_frame, text="Precio del dólar del día")
        frame.pack(fill="x", padx=10, pady=6)

        ttk.Label(frame, text="Precio (Bs):").pack(side=tk.LEFT, padx=6)
        self.entry_dolar = ttk.Entry(frame, width=12)
        self.entry_dolar.pack(side=tk.LEFT)

        ttk.Button(frame, text="Actualizar", style="Primary.TButton", command=self._set_dolar).pack(
            side=tk.LEFT, padx=6
        )

        self.label_proyecto = ttk.Label(frame, text=f"Proyecto: {self.proyecto}", style="Header.TLabel")
        self.label_proyecto.pack(side=tk.RIGHT, padx=6)

    # ============================================
    #         FORMULARIO PRINCIPAL
    # ============================================

    def _crear_formulario(self) -> None:
        self.frame_form = ttk.LabelFrame(self.content_frame, text="Datos del gasto / Mano de obra")
        self.frame_form.pack(fill="x", padx=10, pady=8)

        # Selector de modo
        ttk.Label(self.frame_form, text="Modo:").grid(row=0, column=0, sticky="w", padx=6)
        self.combo_modo = ttk.Combobox(self.frame_form, values=["Factura", "Mano de obra"],
                                       state="readonly", width=18)
        self.combo_modo.grid(row=1, column=0, sticky="w", padx=6)
        self.combo_modo.set("Factura")
        self.combo_modo.bind("<<ComboboxSelected>>", lambda e: self._aplicar_estado_por_modo())

        # -----------------------------
        # CAMPOS PARA FACTURA
        # -----------------------------
        labels_factura = [
            "Factura",
            "Descripción",
            "Unidad",
            "Cantidad",
            "Precio sin IVA (Bs)",
            "Precio con IVA (Bs)",
            "Fecha (DD/MM/YYYY)",
            "Proveedor",
        ]
        self.entries_factura: list[ttk.Entry] = []
        self.labels_factura_widgets: list[ttk.Label] = []

        for col, texto in enumerate(labels_factura, start=1):
            label = ttk.Label(self.frame_form, text=texto)
            label.grid(row=0, column=col, padx=6, pady=4, sticky="w")
            entry = ttk.Entry(self.frame_form, width=20)
            entry.grid(row=1, column=col, padx=6, pady=2)

            self.entries_factura.append(entry)
            self.labels_factura_widgets.append(label)

        (
            self.entry_factura,
            self.entry_desc,
            self.entry_unidad,
            self.entry_cantidad,
            self.entry_precio_sin_iva,
            self.entry_precio_con_iva,
            self.entry_fecha,
            self.entry_proveedor,
        ) = self.entries_factura

        # -----------------------------
        # CAMPOS PARA MANO DE OBRA
        # -----------------------------
        self.labels_mano = []

        label_semana = ttk.Label(self.frame_form, text="Semana")
        label_semana.grid(row=0, column=9, padx=6, pady=4, sticky="w")
        self.entry_semana = ttk.Entry(self.frame_form, width=18)
        self.entry_semana.grid(row=1, column=9, padx=6, pady=2)

        label_manoobra = ttk.Label(self.frame_form, text="Monto manoobra ($)")
        label_manoobra.grid(row=0, column=10, padx=6, pady=4, sticky="w")
        self.entry_monto_mano_usd = ttk.Entry(self.frame_form, width=18)
        self.entry_monto_mano_usd.grid(row=1, column=10, padx=6, pady=2)

        label_extras = ttk.Label(self.frame_form, text="Gastos extras ($)")
        label_extras.grid(row=0, column=11, padx=6, pady=4, sticky="w")
        self.entry_gastos_extras_usd = ttk.Entry(self.frame_form, width=18)
        self.entry_gastos_extras_usd.grid(row=1, column=11, padx=6, pady=2)

        self.labels_mano.extend([label_semana, label_manoobra, label_extras])

        # Aplicar visibilidad inicial
        self._aplicar_estado_por_modo()

        # -----------------------------
        # BOTONES DE ACCIÓN
        # -----------------------------
        botones = ttk.Frame(self.content_frame)
        botones.pack(fill="x", padx=10, pady=6)

        self.btn_agregar = ttk.Button(botones, text="Agregar gasto",
                                      style="Success.TButton",
                                      command=self._agregar_o_actualizar)
        self.btn_agregar.pack(side=tk.LEFT, padx=4)

        ttk.Button(botones, text="Cancelar edición",
                   command=self._cancelar_edicion).pack(side=tk.LEFT, padx=4)

        ttk.Button(botones, text="Eliminar seleccionado(s)",
                   style="Danger.TButton",
                   command=self._eliminar_gasto).pack(side=tk.LEFT, padx=4)

        ttk.Button(botones, text="Limpiar lista",
                   command=self._limpiar_lista).pack(side=tk.LEFT, padx=4)

        # -------------- NUEVO: PRESUPUESTOS + EXPORTACIÓN --------------
        export_frame = ttk.Frame(botones)
        export_frame.pack(side=tk.RIGHT)

        # Presupuesto para facturas
        ttk.Label(export_frame, text="Presupuesto Facturas ($):").pack(side=tk.LEFT, padx=4)
        self.entry_presupuesto_facturas = ttk.Entry(export_frame, width=12)
        self.entry_presupuesto_facturas.pack(side=tk.LEFT, padx=4)

        # Presupuesto para mano de obra
        ttk.Label(export_frame, text="Presupuesto Mano Obra ($):").pack(side=tk.LEFT, padx=4)
        self.entry_presupuesto_mano = ttk.Entry(export_frame, width=12)
        self.entry_presupuesto_mano.pack(side=tk.LEFT, padx=4)

        # Botón exportar SOLO FACTURAS
        ttk.Button(
            export_frame,
            text="Exportar Facturas",
            style="Primary.TButton",
            command=lambda: self._exportar_excel(tipo="Factura"),
        ).pack(side=tk.LEFT, padx=4)

        # Botón exportar SOLO MANO DE OBRA
        ttk.Button(
            export_frame,
            text="Exportar Mano Obra",
            style="Primary.TButton",
            command=lambda: self._exportar_excel(tipo="Mano de obra"),
        ).pack(side=tk.LEFT, padx=4)

    # ============================================
    #   CONTROL DE VISIBILIDAD SEGÚN MODO
    # ============================================

    def _aplicar_estado_por_modo(self) -> None:
        modo = self.combo_modo.get()
        es_mano = modo == "Mano de obra"

        # Factura: ocultar si es mano de obra
        for label, entry in zip(self.labels_factura_widgets, self.entries_factura):
            if es_mano:
                label.grid_remove()
                entry.grid_remove()
            else:
                label.grid()
                entry.grid()

        # Mano de obra: ocultar si es factura
        for label, entry in zip(
            self.labels_mano,
            [self.entry_semana, self.entry_monto_mano_usd, self.entry_gastos_extras_usd],
        ):
            if es_mano:
                label.grid()
                entry.grid()
            else:
                label.grid_remove()
                entry.grid_remove()

    # ============================================
    #             CONTROLES DE TABLA
    # ============================================

    def _crear_controles_tabla(self) -> None:
        filtros = ttk.LabelFrame(self.content_frame, text="Búsqueda y filtros")
        filtros.pack(fill="x", padx=10, pady=8)

        ttk.Label(filtros, text="Buscar:").grid(row=0, column=0, padx=4, sticky="w")
        self.entry_buscar = ttk.Entry(filtros, width=30)
        self.entry_buscar.grid(row=1, column=0, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Proveedor:").grid(row=0, column=1, padx=4, sticky="w")
        self.entry_filtro_proveedor = ttk.Entry(filtros, width=20)
        self.entry_filtro_proveedor.grid(row=1, column=1, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Fecha desde (DD/MM/YYYY):").grid(row=0, column=2, padx=4, sticky="w")
        self.entry_fecha_desde = ttk.Entry(filtros, width=15)
        self.entry_fecha_desde.grid(row=1, column=2, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Fecha hasta (DD/MM/YYYY):").grid(row=0, column=3, padx=4, sticky="w")
        self.entry_fecha_hasta = ttk.Entry(filtros, width=15)
        self.entry_fecha_hasta.grid(row=1, column=3, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Monto mínimo (Bs):").grid(row=0, column=4, padx=4, sticky="w")
        self.entry_monto_min = ttk.Entry(filtros, width=12)
        self.entry_monto_min.grid(row=1, column=4, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Monto máximo (Bs):").grid(row=0, column=5, padx=4, sticky="w")
        self.entry_monto_max = ttk.Entry(filtros, width=12)
        self.entry_monto_max.grid(row=1, column=5, padx=4, pady=2, sticky="w")

        ttk.Button(filtros, text="Aplicar filtros", command=self._aplicar_filtros).grid(row=1, column=6, padx=4)
        ttk.Button(filtros, text="Limpiar filtros", command=self._limpiar_filtros).grid(row=1, column=7, padx=4)

    # ============================================
    #                 TABLA
    # ============================================

    def _crear_tabla(self) -> None:
        frame = ttk.Frame(self.content_frame)
        frame.pack(fill="both", expand=True, padx=10, pady=8)

        self.tabla = ttk.Treeview(frame, columns=COLUMNAS, show="headings", height=15)

        for col in COLUMNAS:
            self.tabla.heading(col, text=col)
            ancho = 120 if col not in ("Descripción", "Precio sin IVA (Bs)", "Precio con IVA (Bs)") else 180
            self.tabla.column(col, width=ancho, anchor="center")

        scroll_y = ttk.Scrollbar(frame, orient="vertical", command=self.tabla.yview)
        scroll_x = ttk.Scrollbar(frame, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(yscroll=scroll_y.set, xscroll=scroll_x.set)

        self.tabla.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tabla.bind("<Double-1>", self._iniciar_edicion)

    # ============================================
    #              RESUMEN
    # ============================================

    def _crear_resumen(self) -> None:
        frame = ttk.Frame(self.content_frame)
        frame.pack(fill="x", padx=10, pady=6)

        self.label_total_bs = ttk.Label(frame, text="Total Bs: 0.00", style="Header.TLabel")
        self.label_total_bs.pack(side=tk.LEFT, padx=8)

        self.label_total_usd = ttk.Label(frame, text="Total $: 0.00", style="Header.TLabel")
        self.label_total_usd.pack(side=tk.LEFT, padx=8)

        self.label_cantidad = ttk.Label(frame, text="Registros: 0", style="Header.TLabel")
        self.label_cantidad.pack(side=tk.LEFT, padx=8)

    def _crear_exportacion(self) -> None:
        pass

    @staticmethod
    def _validar_fecha(fecha: str) -> bool:
        try:
            dt.datetime.strptime(fecha, "%d/%m/%Y")
            return True
        except ValueError:
            return False

    def _mostrar_error(self, mensaje: str) -> None:
        messagebox.showerror("Error", mensaje)

    def _limpiar_campos(self) -> None:
        for entry in [
            *self.entries_factura,
            self.entry_semana,
            self.entry_monto_mano_usd,
            self.entry_gastos_extras_usd,
        ]:
            entry.config(state="normal")
            entry.delete(0, tk.END)

        self.registro_en_edicion = None
        self.combo_modo.set("Factura")
        self._aplicar_estado_por_modo()
        self.btn_agregar.config(text="Agregar gasto")

    def _cancelar_edicion(self) -> None:
        self._limpiar_campos()

    def _limpiar_lista(self) -> None:
        if not self.registros:
            return

        if messagebox.askyesno(
            "Confirmar",
            "Esto eliminará todos los registros de este proyecto. ¿Deseas continuar?",
        ):
            self.registros.clear()
            self._guardar_datos()
            self._limpiar_campos()
            self._refrescar_tabla()

    def _recalcular_items(self) -> None:
        for idx, registro in enumerate(self.registros, start=1):
            registro.item = idx
    # ============================================
    #                 FILTROS
    # ============================================

    def _parse_fecha(self, fecha: str) -> Optional[dt.datetime]:
        try:
            return dt.datetime.strptime(fecha, "%d/%m/%Y")
        except:
            return None

    def _limpiar_filtros(self) -> None:
        for entry in [
            self.entry_buscar,
            self.entry_filtro_proveedor,
            self.entry_fecha_desde,
            self.entry_fecha_hasta,
            self.entry_monto_min,
            self.entry_monto_max,
        ]:
            entry.delete(0, tk.END)
        self._refrescar_tabla()

    def _aplicar_filtros(self) -> None:
        texto = self.entry_buscar.get().lower().strip()
        proveedor = self.entry_filtro_proveedor.get().lower().strip()

        fecha_desde = self._parse_fecha(self.entry_fecha_desde.get())
        fecha_hasta = self._parse_fecha(self.entry_fecha_hasta.get())

        try:
            monto_min = float(self.entry_monto_min.get()) if self.entry_monto_min.get() else None
            monto_max = float(self.entry_monto_max.get()) if self.entry_monto_max.get() else None
        except:
            self._mostrar_error("Los montos deben ser numéricos.")
            return

        resultados = []
        for r in self.registros:
            if texto:
                cad = f"{r.factura} {r.descripcion} {r.proveedor}".lower()
                if texto not in cad:
                    continue

            if proveedor and proveedor not in r.proveedor.lower():
                continue

            f = self._parse_fecha(r.fecha) if r.fecha else None
            if fecha_desde and (not f or f < fecha_desde):
                continue
            if fecha_hasta and (not f or f > fecha_hasta):
                continue

            if monto_min is not None and r.total_sin_iva < monto_min:
                continue
            if monto_max is not None and r.total_sin_iva > monto_max:
                continue

            resultados.append(r)

        self._refrescar_tabla(resultados)

    # ============================================
    #           ACTUALIZAR TABLA Y RESUMEN
    # ============================================

    def _refrescar_tabla(self, registros=None) -> None:
        registros = registros if registros is not None else self.registros

        for row in self.tabla.get_children():
            self.tabla.delete(row)

        for r in registros:
            self.tabla.insert("", "end", values=r.to_row())

        self._actualizar_resumen(registros)

    def _actualizar_resumen(self, registros) -> None:
        total_bs = sum(r.precio_con_iva_bs * (r.cantidad if r.cantidad else 1) for r in registros)

        total_usd = round(total_bs / self.precio_dolar, 2) if self.precio_dolar else 0

        self.label_total_bs.config(text=f"Total Bs (c/IVA): {total_bs:,.2f}")
        self.label_total_usd.config(text=f"Total $ (c/IVA): {total_usd:,.2f}")
        self.label_cantidad.config(text=f"Registros: {len(registros)}")
        self.label_proyecto.config(text=f"Proyecto: {self.proyecto}")

    # ============================================
    #              GUARDAR Y CARGAR
    # ============================================

    def _leer_archivo_datos(self) -> pd.DataFrame:
        if not os.path.exists(ARCHIVO_DATOS):
            return pd.DataFrame()
        try:
            return pd.read_excel(ARCHIVO_DATOS)
        except:
            return pd.DataFrame()

    def _cargar_datos_proyecto(self) -> None:
        self.registros.clear()

        if self._df_global.empty:
            return

        if "Proyecto" in self._df_global.columns:
            df_proj = self._df_global[self._df_global["Proyecto"] == self.proyecto]
        else:
            df_proj = pd.DataFrame()

        if df_proj.empty:
            self._recalcular_items()
            return

        for _, fila in df_proj.iterrows():
            registro = RegistroGasto.from_dict(fila.to_dict())
            self.registros.append(registro)

        self._recalcular_items()

    def _guardar_datos(self) -> None:
        df_nuevo = pd.DataFrame([asdict(r) for r in self.registros])

        df_nuevo.rename(columns={
            "item": "Item",
            "tipo": "Tipo",
            "factura": "Factura",
            "descripcion": "Descripción",
            "unidad": "Unidad",
            "cantidad": "Cantidad",
            "precio_sin_iva_bs": "Precio sin IVA (Bs)",
            "precio_con_iva_bs": "Precio con IVA (Bs)",
            "precio_sin_iva_usd": "Precio sin IVA ($)",
            "precio_con_iva_usd": "Precio con IVA ($)",
            "fecha": "Fecha",
            "total_sin_iva": "Total sin IVA (Bs)",
            "proveedor": "Proveedor",
            "proyecto": "Proyecto",
            "semana": "Semana",
            "monto_manoobra_usd": "Monto Manoobra ($)",
            "gastos_extras_usd": "Gastos extras ($)",
        }, inplace=True)

        df_global = self._leer_archivo_datos()

        # Reemplazar solo el proyecto actual
        if not df_global.empty and "Proyecto" in df_global.columns:
            df_sin_actual = df_global[df_global["Proyecto"] != self.proyecto]
            df_final = pd.concat([df_sin_actual, df_nuevo], ignore_index=True)
        else:
            df_final = df_nuevo

        try:
            df_final.to_excel(ARCHIVO_DATOS, index=False)
            self._df_global = df_final
        except Exception as e:
            self._mostrar_error(f"No se pudo guardar el archivo: {e}")

    def _set_dolar(self) -> None:
        try:
            self.precio_dolar = float(self.entry_dolar.get())
        except ValueError:
            self._mostrar_error("Ingresa un valor numérico válido para el dólar.")
            return

        messagebox.showinfo("Éxito", f"Precio del dólar actualizado a Bs. {self.precio_dolar}")

        for registro in self.registros:
            if registro.precio_sin_iva_bs:
                registro.precio_sin_iva_usd = (
                    round(registro.precio_sin_iva_bs / self.precio_dolar, 2)
                    if self.precio_dolar > 0
                    else 0.0
                )

            if registro.precio_con_iva_bs:
                registro.precio_con_iva_usd = (
                    round(registro.precio_con_iva_bs / self.precio_dolar, 2)
                    if self.precio_dolar > 0
                    else 0.0
                )

        self._refrescar_tabla()

    def _agregar_o_actualizar(self) -> None:
        modo = self.combo_modo.get()

        if not self.entry_desc.get().strip():
            self._mostrar_error("La descripción es obligatoria.")
            return

        fecha_texto = self.entry_fecha.get().strip()
        if fecha_texto and not self._validar_fecha(fecha_texto):
            self._mostrar_error("Formato de fecha inválido. Use DD/MM/YYYY")
            return

        if modo == "Factura":
            if self.precio_dolar <= 0:
                messagebox.showwarning("Atención", "Primero ingresa el precio del dólar del día.")
                return

            campos = [
                self.entry_factura,
                self.entry_unidad,
                self.entry_cantidad,
                self.entry_precio_sin_iva,
                self.entry_precio_con_iva,
                self.entry_fecha,
                self.entry_proveedor,
            ]

            if any(not c.get().strip() for c in campos):
                self._mostrar_error("Todos los campos de factura son obligatorios.")
                return

            try:
                cantidad = float(self.entry_cantidad.get())
                precio_sin = float(self.entry_precio_sin_iva.get())
                precio_con = float(self.entry_precio_con_iva.get())
            except ValueError:
                self._mostrar_error("Verifica números (cantidad/precios).")
                return

            precio_sin_usd = round(precio_sin / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0
            precio_con_usd = round(precio_con / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0
            total_sin = round(precio_sin * cantidad, 2)

            registro = RegistroGasto(
                item=len(self.registros) + 1,
                tipo="Factura",
                factura=self.entry_factura.get(),
                descripcion=self.entry_desc.get(),
                unidad=self.entry_unidad.get(),
                cantidad=cantidad,
                precio_sin_iva_bs=precio_sin,
                precio_con_iva_bs=precio_con,
                precio_sin_iva_usd=precio_sin_usd,
                precio_con_iva_usd=precio_con_usd,
                fecha=fecha_texto,
                total_sin_iva=total_sin,
                proveedor=self.entry_proveedor.get(),
                proyecto=self.proyecto,
            )
        else:
            if not self.entry_semana.get().strip():
                self._mostrar_error("La semana es obligatoria para mano de obra.")
                return

            try:
                monto_mano = float(self.entry_monto_mano_usd.get() or 0)
                gastos_extra = float(self.entry_gastos_extras_usd.get() or 0)
            except ValueError:
                self._mostrar_error("Verifica números en monto de mano de obra y gastos extras.")
                return

            registro = RegistroGasto(
                item=len(self.registros) + 1,
                tipo="Mano de obra",
                factura=self.entry_factura.get(),
                descripcion=self.entry_desc.get(),
                unidad=self.entry_unidad.get(),
                cantidad=float(self.entry_cantidad.get() or 0),
                precio_sin_iva_bs=0,
                precio_con_iva_bs=0,
                precio_sin_iva_usd=0,
                precio_con_iva_usd=0,
                fecha=fecha_texto,
                total_sin_iva=0,
                proveedor=self.entry_proveedor.get(),
                proyecto=self.proyecto,
                semana=self.entry_semana.get(),
                monto_manoobra_usd=monto_mano,
                gastos_extras_usd=gastos_extra,
            )

        if self.registro_en_edicion is not None:
            registro.item = self.registro_en_edicion + 1
            self.registros[self.registro_en_edicion] = registro
        else:
            self.registros.append(registro)

        self._recalcular_items()
        self._guardar_datos()
        self._limpiar_campos()
        self._refrescar_tabla()

    def _eliminar_gasto(self) -> None:
        seleccion = self.tabla.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Selecciona al menos un registro para eliminar.")
            return

        indices = [self.tabla.index(i) for i in seleccion]

        if messagebox.askyesno("Confirmar", "¿Seguro que deseas eliminar los registros seleccionados?"):
            for idx in sorted(indices, reverse=True):
                try:
                    self.registros.pop(idx)
                except IndexError:
                    continue

            self._recalcular_items()
            self._guardar_datos()
            self._refrescar_tabla()

    def _iniciar_edicion(self, _event) -> None:
        item_sel = self.tabla.selection()
        if not item_sel:
            return

        idx = self.tabla.index(item_sel[0])
        try:
            registro = self.registros[idx]
        except IndexError:
            return

        self.registro_en_edicion = idx
        self.combo_modo.set(registro.tipo)
        self._aplicar_estado_por_modo()

        self.entry_factura.delete(0, tk.END)
        self.entry_factura.insert(0, registro.factura)

        self.entry_desc.delete(0, tk.END)
        self.entry_desc.insert(0, registro.descripcion)

        self.entry_unidad.delete(0, tk.END)
        self.entry_unidad.insert(0, registro.unidad)

        self.entry_cantidad.delete(0, tk.END)
        self.entry_cantidad.insert(0, str(registro.cantidad))

        self.entry_precio_sin_iva.delete(0, tk.END)
        self.entry_precio_sin_iva.insert(0, str(registro.precio_sin_iva_bs))

        self.entry_precio_con_iva.delete(0, tk.END)
        self.entry_precio_con_iva.insert(0, str(registro.precio_con_iva_bs))

        self.entry_fecha.delete(0, tk.END)
        self.entry_fecha.insert(0, registro.fecha)

        self.entry_proveedor.delete(0, tk.END)
        self.entry_proveedor.insert(0, registro.proveedor)

        self.entry_semana.delete(0, tk.END)
        self.entry_semana.insert(0, registro.semana)

        self.entry_monto_mano_usd.delete(0, tk.END)
        self.entry_monto_mano_usd.insert(0, str(registro.monto_manoobra_usd))

        self.entry_gastos_extras_usd.delete(0, tk.END)
        self.entry_gastos_extras_usd.insert(0, str(registro.gastos_extras_usd))

        self.btn_agregar.config(text="Actualizar registro")

    # ============================================
    #             EXPORTACIÓN A EXCEL
    # ============================================

    def _exportar_excel(self, tipo: str) -> None:
        if not self.registros:
            messagebox.showwarning("Atención", "No hay registros para exportar.")
            return

        # Filtrar registros según el tipo solicitado
        registros_filtrados = [r for r in self.registros if r.tipo == tipo]

        if not registros_filtrados:
            messagebox.showwarning("Atención", f"No hay registros del tipo '{tipo}' para exportar.")
            return

        # Selección del presupuesto según tipo
        if tipo == "Factura":
            try:
                presupuesto = float(self.entry_presupuesto_facturas.get())
            except:
                self._mostrar_error("Ingresa un presupuesto válido para FACTURAS.")
                return
        else:
            try:
                presupuesto = float(self.entry_presupuesto_mano.get())
            except:
                self._mostrar_error("Ingresa un presupuesto válido para MANO DE OBRA.")
                return

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")],
            title="Guardar reporte",
            initialfile=f"reporte_{tipo}_{self.proyecto}.xlsx",
        )

        if not archivo:
            return

        wb = Workbook()
        ws = wb.active

        # ------------------- TÍTULO -------------------
        titulo = f"REPORTE DE GASTOS - {self.proyecto} - {tipo.upper()}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
        cell = ws.cell(row=1, column=1, value=titulo)
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal="center")

        # ------------------- ENCABEZADOS -------------------
        header_row = 3
        for col_idx, encabezado in enumerate(COLUMNAS, start=1):
            c = ws.cell(row=header_row, column=col_idx, value=encabezado)
            c.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")

        # ------------------- FILAS -------------------
        for row_idx, registro in enumerate(registros_filtrados, start=header_row + 1):
            for col_idx, valor in enumerate(registro.to_row(), start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=valor)
                c.alignment = Alignment(horizontal="center")

        # ------------------- TOTALES -------------------
        total_sin_iva_bs = sum(r.precio_sin_iva_bs * (r.cantidad if r.cantidad else 1) for r in registros_filtrados)
        total_con_iva_bs = sum(r.precio_con_iva_bs * (r.cantidad if r.cantidad else 1) for r in registros_filtrados)

        total_con_iva_usd = round(total_con_iva_bs / self.precio_dolar, 2) if self.precio_dolar else 0

        fila_total = header_row + len(registros_filtrados) + 2

        ws.cell(row=fila_total, column=1, value="PRESUPUESTO ($):")
        ws.cell(row=fila_total, column=2, value=presupuesto)

        ws.cell(row=fila_total + 1, column=1, value="TOTAL SIN IVA (Bs):")
        ws.cell(row=fila_total + 1, column=2, value=round(total_sin_iva_bs, 2))

        ws.cell(row=fila_total + 2, column=1, value="TOTAL CON IVA (Bs):")
        ws.cell(row=fila_total + 2, column=2, value=round(total_con_iva_bs, 2))

        utilidad = round(presupuesto - total_con_iva_usd, 2)
        porc = (utilidad / presupuesto * 100) if presupuesto else 0

        ws.cell(row=fila_total + 3, column=1, value="UTILIDAD ($):")
        ws.cell(row=fila_total + 3, column=2, value=utilidad)

        ws.cell(row=fila_total + 4, column=1, value="% UTILIDAD:")
        ws.cell(row=fila_total + 4, column=2, value=round(porc, 2))

        # Colorear totales
        for i in range(fila_total, fila_total + 5):
            for j in range(1, 3):
                c = ws.cell(row=i, column=j)
                c.fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")

        # Ajuste de columnas
        for idx, col in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 3

        wb.save(archivo)
        wb.close()
        messagebox.showinfo("Éxito", f"Reporte '{tipo}' generado correctamente.")

    # ============================================
    #     SELECCIONAR PROYECTO
    # ============================================

    def _seleccionar_proyecto_modal(self) -> None:
        proyectos = []
        if hasattr(self, "_df_global") and not self._df_global.empty and "Proyecto" in self._df_global.columns:
            proyectos = sorted(self._df_global["Proyecto"].dropna().unique().tolist())

        modal = tk.Toplevel(self.root)
        modal.title("Seleccionar proyecto")
        modal.geometry("360x150")
        modal.transient(self.root)
        modal.grab_set()

        ttk.Label(modal, text="Proyecto existente:").pack(anchor="w", padx=10, pady=4)
        combo = ttk.Combobox(modal, values=proyectos, width=35, state="readonly" if proyectos else "normal")
        combo.pack(padx=10, pady=4)
        if proyectos:
            combo.set(proyectos[0])

        ttk.Label(modal, text="O crear uno nuevo:").pack(anchor="w", padx=10, pady=4)
        entry_new = ttk.Entry(modal, width=37)
        entry_new.pack(padx=10, pady=4)

        def aceptar():
            elegido = entry_new.get().strip() or combo.get().strip()
            if not elegido:
                self._mostrar_error("Debes seleccionar o crear un proyecto.")
                return
            self.proyecto = elegido
            modal.destroy()

        ttk.Button(modal, text="Aceptar", style="Primary.TButton", command=aceptar).pack(pady=8)
        self.root.wait_window(modal)

    # ============================================
    #              EJECUCIÓN
    # ============================================

    def ejecutar(self) -> None:
        self.root.mainloop()


# =========================
#   PUNTO DE ENTRADA
# =========================

if __name__ == "__main__":
    app = GestorGastos()
    app.ejecutar()
