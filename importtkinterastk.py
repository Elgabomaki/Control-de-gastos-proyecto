"""Aplicación de control de gastos con Tkinter.

Soporte para:
- Registro de facturas
- Registro de mano de obra (modo alternativo en el mismo formulario)
- Selección de proyecto al iniciar
- Exportación con totales y presupuesto en USD

Comportamiento solicitado:
- El formulario cambia dinámicamente según el "Modo" seleccionado en la UI (Factura / Mano de obra).
  No se guarda un "tipo por defecto" por proyecto; el usuario elige el modo en el formulario.
"""

from __future__ import annotations

import datetime as dt
import os
import tkinter as tk
from dataclasses import dataclass, asdict
from tkinter import filedialog, messagebox, ttk
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ARCHIVO_DATOS = "gastos_guardados.xlsx"
IVA = 0.16

COLUMNAS = [
    "Item",
    "Tipo",
    "Factura",
    "Descripción",
    "Unidad",
    "Cantidad",
    "Precio sin IVA (Bs)",
    "Precio con IVA (Bs)",
    "Precio sin IVA ($)",
    "Precio con IVA ($)",
    "Fecha",
    "Total sin IVA (Bs)",
    "Proveedor",
    "Semana",
    "Monto Manoobra ($)",
    "Gastos extras ($)",
]


@dataclass
class RegistroGasto:
    item: int
    tipo: str  # "Factura" o "Mano de obra"
    factura: str
    descripcion: str
    unidad: str
    cantidad: float
    precio_sin_iva_bs: float
    precio_con_iva_bs: float
    precio_sin_iva_usd: float
    precio_con_iva_usd: float
    fecha: str
    total_sin_iva: float
    proveedor: str
    proyecto: str
    semana: str = ""
    monto_manoobra_usd: float = 0.0
    gastos_extras_usd: float = 0.0

    @classmethod
    def from_dict(cls, data: dict[str, object]) -> "RegistroGasto":
        precio_sin_bs = float(data.get("Precio sin IVA (Bs)") or data.get("Precio (Bs)") or 0)
        precio_con_bs = float(data.get("Precio con IVA (Bs)") or round(precio_sin_bs * (1 + IVA), 2))
        precio_sin_usd = float(data.get("Precio sin IVA ($)") or data.get("Precio ($)") or 0)
        precio_con_usd = float(data.get("Precio con IVA ($)") or data.get("Precio ($)") or 0)
        cantidad = float(data.get("Cantidad", 0) or 0)
        total_sin = float(data.get("Total sin IVA (Bs)") or data.get("Total") or precio_sin_bs * cantidad)
        return cls(
            item=int(data.get("Item", 0) or 0),
            tipo=str(data.get("Tipo", "Factura") or "Factura"),
            factura=str(data.get("Factura", "") or ""),
            descripcion=str(data.get("Descripción", "") or ""),
            unidad=str(data.get("Unidad", "") or ""),
            cantidad=cantidad,
            precio_sin_iva_bs=precio_sin_bs,
            precio_con_iva_bs=precio_con_bs,
            precio_sin_iva_usd=precio_sin_usd,
            precio_con_iva_usd=precio_con_usd,
            fecha=str(data.get("Fecha", "") or ""),
            total_sin_iva=total_sin,
            proveedor=str(data.get("Proveedor", "") or ""),
            proyecto=str(data.get("Proyecto", "") or ""),
            semana=str(data.get("Semana", "") or ""),
            monto_manoobra_usd=float(data.get("Monto Manoobra ($)", 0) or 0),
            gastos_extras_usd=float(data.get("Gastos extras ($)", 0) or 0),
        )

    def to_row(self) -> List[object]:
        return [
            self.item,
            self.tipo,
            self.factura,
            self.descripcion,
            self.unidad,
            self.cantidad,
            self.precio_sin_iva_bs,
            self.precio_con_iva_bs,
            self.precio_sin_iva_usd,
            self.precio_con_iva_usd,
            self.fecha,
            self.total_sin_iva,
            self.proveedor,
            self.semana,
            self.monto_manoobra_usd,
            self.gastos_extras_usd,
        ]


class GestorGastos:
    def __init__(self) -> None:
        self.registros: List[RegistroGasto] = []
        self.precio_dolar: float = 0.0
        self.registro_en_edicion: Optional[int] = None
        self.proyecto: str = ""

        self.root = tk.Tk()
        self.root.title("Control de Gastos - Terra Caliza")
        self.root.geometry("1280x740")
        self.root.resizable(True, True)

        self._configurar_estilos()

        self._df_global = self._leer_archivo_datos()

        # seleccionar proyecto antes de crear la UI completa
        self._seleccionar_proyecto_modal()

        self._crear_componentes()
        self._cargar_datos_proyecto()
        self._refrescar_tabla()

    def _configurar_estilos(self) -> None:
        style = ttk.Style(self.root)
        style.theme_use("clam")
        style.configure("TButton", padding=6, font=("Arial", 10))
        style.configure("Primary.TButton", background="#1976D2", foreground="white")
        style.map("Primary.TButton", background=[("active", "#1565C0")])
        style.configure("Danger.TButton", background="#C62828", foreground="white")
        style.map("Danger.TButton", background=[("active", "#B71C1C")])
        style.configure("Success.TButton", background="#2E7D32", foreground="white")
        style.map("Success.TButton", background=[("active", "#1B5E20")])
        style.configure("TLabel", font=("Arial", 10))
        style.configure("Header.TLabel", font=("Arial", 12, "bold"))

    def _crear_componentes(self) -> None:
        self._crear_frame_dolar()
        self._crear_formulario()       # aquí el combo de modo (Factura / Mano de obra)
        self._crear_controles_tabla()
        self._crear_tabla()
        self._crear_resumen()
        self._crear_exportacion()
        self._crear_menu()

    def _crear_menu(self) -> None:
        menubar = tk.Menu(self.root)
        proyecto_menu = tk.Menu(menubar, tearoff=0)
        proyecto_menu.add_command(label="Cambiar proyecto", command=self._seleccionar_proyecto_modal)
        menubar.add_cascade(label="Proyecto", menu=proyecto_menu)
        self.root.config(menu=menubar)

    def _crear_frame_dolar(self) -> None:
        frame = ttk.LabelFrame(self.root, text="Precio del dólar del día")
        frame.pack(fill="x", padx=10, pady=6)

        ttk.Label(frame, text="Precio (Bs):").pack(side=tk.LEFT, padx=6)
        self.entry_dolar = ttk.Entry(frame, width=12)
        self.entry_dolar.pack(side=tk.LEFT)
        ttk.Button(frame, text="Actualizar", style="Primary.TButton", command=self._set_dolar).pack(
            side=tk.LEFT, padx=6
        )

        self.label_proyecto = ttk.Label(frame, text=f"Proyecto: {self.proyecto}", style="Header.TLabel")
        self.label_proyecto.pack(side=tk.RIGHT, padx=6)

    def _crear_formulario(self) -> None:
        self.frame_form = ttk.LabelFrame(self.root, text="Datos del gasto / Mano de obra")
        self.frame_form.pack(fill="x", padx=10, pady=8)

        # Modo selector en la UI principal (el usuario cambia aquí y el formulario se adapta)
        ttk.Label(self.frame_form, text="Modo:").grid(row=0, column=0, sticky="w", padx=6)
        self.combo_modo = ttk.Combobox(self.frame_form, values=["Factura", "Mano de obra"], state="readonly", width=18)
        self.combo_modo.grid(row=1, column=0, sticky="w", padx=6)
        self.combo_modo.set("Factura")
        self.combo_modo.bind("<<ComboboxSelected>>", lambda e: self._on_modo_cambiado())

        # Campos para Factura (columnas 1..8)
        labels_factura = [
            "Factura",
            "Descripción",
            "Unidad",
            "Cantidad",
            "Precio sin IVA (Bs)",
            "Precio con IVA (Bs)",
            "Fecha (DD/MM/YYYY)",
            "Proveedor",
        ]
        self.entries_factura: list[ttk.Entry] = []
        for col, texto in enumerate(labels_factura, start=1):
            ttk.Label(self.frame_form, text=texto).grid(row=0, column=col, padx=6, pady=4, sticky="w")
            entry = ttk.Entry(self.frame_form, width=20)
            entry.grid(row=1, column=col, padx=6, pady=2)
            self.entries_factura.append(entry)

        (
            self.entry_factura,
            self.entry_desc,
            self.entry_unidad,
            self.entry_cantidad,
            self.entry_precio_sin_iva,
            self.entry_precio_con_iva,
            self.entry_fecha,
            self.entry_proveedor,
        ) = self.entries_factura

        # Campos para Mano de obra (columnas 9..11) — visibles solo en ese modo
        ttk.Label(self.frame_form, text="Semana").grid(row=0, column=9, padx=6, pady=4, sticky="w")
        self.entry_semana = ttk.Entry(self.frame_form, width=18)
        self.entry_semana.grid(row=1, column=9, padx=6, pady=2)

        ttk.Label(self.frame_form, text="Monto manoobra ($)").grid(row=0, column=10, padx=6, pady=4, sticky="w")
        self.entry_monto_mano_usd = ttk.Entry(self.frame_form, width=18)
        self.entry_monto_mano_usd.grid(row=1, column=10, padx=6, pady=2)

        ttk.Label(self.frame_form, text="Gastos extras ($)").grid(row=0, column=11, padx=6, pady=4, sticky="w")
        self.entry_gastos_extras_usd = ttk.Entry(self.frame_form, width=18)
        self.entry_gastos_extras_usd.grid(row=1, column=11, padx=6, pady=2)

        # aplicar estado inicial (Factura)
        self._aplicar_estado_por_modo()

        # Botones acción
        botones = ttk.Frame(self.root)
        botones.pack(fill="x", padx=10, pady=6)
        self.btn_agregar = ttk.Button(botones, text="Agregar gasto", style="Success.TButton", command=self._agregar_o_actualizar)
        self.btn_agregar.pack(side=tk.LEFT, padx=4)
        ttk.Button(botones, text="Cancelar edición", command=self._cancelar_edicion).pack(side=tk.LEFT, padx=4)
        ttk.Button(botones, text="Eliminar seleccionado(s)", style="Danger.TButton", command=self._eliminar_gasto).pack(
            side=tk.LEFT, padx=4
        )
        ttk.Button(botones, text="Limpiar lista", command=self._limpiar_lista).pack(side=tk.LEFT, padx=4)

        export_frame = ttk.Frame(botones)
        export_frame.pack(side=tk.RIGHT)
        ttk.Label(export_frame, text="Presupuesto ($):").pack(side=tk.LEFT, padx=4)
        self.entry_presupuesto = ttk.Entry(export_frame, width=14)
        self.entry_presupuesto.pack(side=tk.LEFT, padx=4)
        ttk.Button(export_frame, text="Exportar a Excel", style="Primary.TButton", command=self._exportar_excel).pack(
            side=tk.LEFT, padx=4
        )

    def _on_modo_cambiado(self) -> None:
        """Handler llamado cuando el usuario cambia el modo en el combobox."""
        self._aplicar_estado_por_modo()

    def _aplicar_estado_por_modo(self) -> None:
        """Habilita / deshabilita campos según el modo actual."""
        modo = self.combo_modo.get() if hasattr(self, "combo_modo") else "Factura"
        es_mano = modo == "Mano de obra"

        # Campos factura activados solo si no es mano
        estado_factura = "normal" if not es_mano else "disabled"
        for e in self.entries_factura:
            try:
                e.config(state=estado_factura)
                if es_mano:
                    # opcional: limpiar campos factura al cambiar a mano de obra
                    e.delete(0, tk.END)
            except Exception:
                pass

        # Campos mano de obra
        estado_mano = "normal" if es_mano else "disabled"
        for e in [self.entry_semana, self.entry_monto_mano_usd, self.entry_gastos_extras_usd]:
            try:
                e.config(state=estado_mano)
                if not es_mano:
                    e.delete(0, tk.END)
            except Exception:
                pass

        # Actualizar texto o tooltip si se desea
        try:
            self.label_proyecto.config(text=f"Proyecto: {self.proyecto}")
        except Exception:
            pass

    def _crear_controles_tabla(self) -> None:
        filtros = ttk.LabelFrame(self.root, text="Búsqueda y filtros")
        filtros.pack(fill="x", padx=10, pady=8)

        ttk.Label(filtros, text="Buscar (factura/proveedor/desc.):").grid(row=0, column=0, padx=4, pady=2, sticky="w")
        self.entry_buscar = ttk.Entry(filtros, width=28)
        self.entry_buscar.grid(row=1, column=0, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Proveedor:").grid(row=0, column=1, padx=4, pady=2, sticky="w")
        self.entry_filtro_proveedor = ttk.Entry(filtros, width=20)
        self.entry_filtro_proveedor.grid(row=1, column=1, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Fecha desde (DD/MM/YYYY):").grid(row=0, column=2, padx=4, pady=2, sticky="w")
        self.entry_fecha_desde = ttk.Entry(filtros, width=14)
        self.entry_fecha_desde.grid(row=1, column=2, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Fecha hasta (DD/MM/YYYY):").grid(row=0, column=3, padx=4, pady=2, sticky="w")
        self.entry_fecha_hasta = ttk.Entry(filtros, width=14)
        self.entry_fecha_hasta.grid(row=1, column=3, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Monto mínimo (Bs):").grid(row=0, column=4, padx=4, pady=2, sticky="w")
        self.entry_monto_min = ttk.Entry(filtros, width=12)
        self.entry_monto_min.grid(row=1, column=4, padx=4, pady=2, sticky="w")

        ttk.Label(filtros, text="Monto máximo (Bs):").grid(row=0, column=5, padx=4, pady=2, sticky="w")
        self.entry_monto_max = ttk.Entry(filtros, width=12)
        self.entry_monto_max.grid(row=1, column=5, padx=4, pady=2, sticky="w")

        ttk.Button(filtros, text="Aplicar filtros", command=self._aplicar_filtros).grid(row=1, column=6, padx=6)
        ttk.Button(filtros, text="Limpiar filtros", command=self._limpiar_filtros).grid(row=1, column=7, padx=6)

    def _crear_tabla(self) -> None:
        frame = ttk.Frame(self.root)
        frame.pack(fill="both", expand=True, padx=10, pady=8)

        self.tabla = ttk.Treeview(frame, columns=COLUMNAS, show="headings", height=15)
        for col in COLUMNAS:
            self.tabla.heading(col, text=col)
            ancho = 120 if col not in ("Descripción", "Precio con IVA (Bs)", "Precio sin IVA (Bs)") else 180
            self.tabla.column(col, width=ancho, anchor="center")

        scrollbar_y = ttk.Scrollbar(frame, orient="vertical", command=self.tabla.yview)
        scrollbar_x = ttk.Scrollbar(frame, orient="horizontal", command=self.tabla.xview)
        self.tabla.configure(yscroll=scrollbar_y.set, xscroll=scrollbar_x.set)

        self.tabla.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tabla.bind("<Double-1>", self._iniciar_edicion)

    def _crear_resumen(self) -> None:
        frame = ttk.Frame(self.root)
        frame.pack(fill="x", padx=10, pady=6)

        self.label_total_bs = ttk.Label(frame, text="Total Bs: 0.00", style="Header.TLabel")
        self.label_total_bs.pack(side=tk.LEFT, padx=8)

        self.label_total_usd = ttk.Label(frame, text="Total $: 0.00", style="Header.TLabel")
        self.label_total_usd.pack(side=tk.LEFT, padx=8)

        self.label_cantidad = ttk.Label(frame, text="Registros: 0", style="Header.TLabel")
        self.label_cantidad.pack(side=tk.LEFT, padx=8)

    def _crear_exportacion(self) -> None:
        return

    @staticmethod
    def _validar_fecha(fecha: str) -> bool:
        try:
            dt.datetime.strptime(fecha, "%d/%m/%Y")
            return True
        except ValueError:
            return False

    def _mostrar_error(self, mensaje: str) -> None:
        messagebox.showerror("Error", mensaje)

    def _limpiar_campos(self) -> None:
        for e in self.entries_factura:
            try:
                e.config(state="normal")
                e.delete(0, tk.END)
            except Exception:
                pass
        for e in [self.entry_semana, self.entry_monto_mano_usd, self.entry_gastos_extras_usd]:
            try:
                e.config(state="normal")
                e.delete(0, tk.END)
            except Exception:
                pass
        self.registro_en_edicion = None
        self.combo_modo.set("Factura")
        self._aplicar_estado_por_modo()
        self.btn_agregar.config(text="Agregar gasto")

    def _cancelar_edicion(self) -> None:
        self._limpiar_campos()

    def _limpiar_lista(self) -> None:
        if not self.registros:
            return
        if messagebox.askyesno("Confirmar", "Esto eliminará todos los registros de este proyecto. ¿Deseas continuar?"):
            self.registros.clear()
            self._guardar_datos()
            self._limpiar_campos()
            self._refrescar_tabla()

    def _recalcular_items(self) -> None:
        for idx, registro in enumerate(self.registros, start=1):
            registro.item = idx

    def _leer_archivo_datos(self) -> pd.DataFrame:
        if not os.path.exists(ARCHIVO_DATOS):
            return pd.DataFrame()
        try:
            df = pd.read_excel(ARCHIVO_DATOS)
            return df
        except Exception:
            return pd.DataFrame()

    def _cargar_datos_proyecto(self) -> None:
        self.registros.clear()
        if self._df_global.empty:
            return
        if "Proyecto" in self._df_global.columns:
            try:
                df_proj = self._df_global[self._df_global["Proyecto"] == self.proyecto]
            except Exception:
                df_proj = pd.DataFrame()
        else:
            df_proj = pd.DataFrame()
        if df_proj.empty:
            self._recalcular_items()
            return
        for _, fila in df_proj.iterrows():
            registro = RegistroGasto.from_dict(fila.to_dict())
            self.registros.append(registro)
        self._recalcular_items()

    def _guardar_datos(self) -> None:
        df_nuevo = pd.DataFrame([asdict(r) for r in self.registros])
        df_nuevo.rename(columns={
            "item": "Item",
            "tipo": "Tipo",
            "factura": "Factura",
            "descripcion": "Descripción",
            "unidad": "Unidad",
            "cantidad": "Cantidad",
            "precio_sin_iva_bs": "Precio sin IVA (Bs)",
            "precio_con_iva_bs": "Precio con IVA (Bs)",
            "precio_sin_iva_usd": "Precio sin IVA ($)",
            "precio_con_iva_usd": "Precio con IVA ($)",
            "fecha": "Fecha",
            "total_sin_iva": "Total sin IVA (Bs)",
            "proveedor": "Proveedor",
            "proyecto": "Proyecto",
            "semana": "Semana",
            "monto_manoobra_usd": "Monto Manoobra ($)",
            "gastos_extras_usd": "Gastos extras ($)",
        }, inplace=True)

        df_global = self._leer_archivo_datos()
        if df_global.empty:
            df_final = df_nuevo
        else:
            if "Proyecto" in df_global.columns:
                try:
                    df_sin_proyecto = df_global[df_global["Proyecto"] != self.proyecto]
                except Exception:
                    df_sin_proyecto = df_global.copy()
            else:
                df_sin_proyecto = df_global.copy()
            df_final = pd.concat([df_sin_proyecto, df_nuevo], ignore_index=True, sort=False)
        try:
            df_final.to_excel(ARCHIVO_DATOS, index=False)
            self._df_global = df_final
        except Exception as e:
            messagebox.showerror("Error al guardar", f"No se pudo escribir {ARCHIVO_DATOS}:\n{e}")

    def _set_dolar(self) -> None:
        try:
            self.precio_dolar = float(self.entry_dolar.get())
            messagebox.showinfo("Éxito", f"Precio del dólar actualizado a Bs. {self.precio_dolar}")
            for r in self.registros:
                if r.precio_sin_iva_bs:
                    r.precio_sin_iva_usd = round(r.precio_sin_iva_bs / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0
                if r.precio_con_iva_bs:
                    r.precio_con_iva_usd = round(r.precio_con_iva_bs / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0
            self._refrescar_tabla()
        except ValueError:
            self._mostrar_error("Ingresa un valor numérico válido para el dólar.")

    def _agregar_o_actualizar(self) -> None:
        modo = self.combo_modo.get()
        if not self.entry_desc.get().strip():
            self._mostrar_error("La descripción es obligatoria.")
            return
        # fecha opcional para manoobra pero validamos si hay texto
        fecha_text = self.entry_fecha.get().strip()
        if fecha_text and not self._validar_fecha(fecha_text):
            self._mostrar_error("Formato de fecha inválido. Use DD/MM/YYYY")
            return

        if modo == "Factura":
            if self.precio_dolar <= 0:
                messagebox.showwarning("Atención", "Primero ingresa el precio del dólar del día.")
                return
            campos = [self.entry_factura, self.entry_unidad, self.entry_cantidad, self.entry_precio_sin_iva, self.entry_precio_con_iva, self.entry_fecha, self.entry_proveedor]
            if any(not c.get().strip() for c in campos):
                self._mostrar_error("Todos los campos de factura son obligatorios.")
                return
            try:
                cantidad = float(self.entry_cantidad.get())
                precio_sin = float(self.entry_precio_sin_iva.get())
                precio_con = float(self.entry_precio_con_iva.get())
            except ValueError:
                self._mostrar_error("Verifica números (cantidad/precios).")
                return
            precio_sin_usd = round(precio_sin / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0
            precio_con_usd = round(precio_con / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0
            total_sin = round(precio_sin * cantidad, 2)
            registro = RegistroGasto(
                item=len(self.registros) + 1,
                tipo="Factura",
                factura=self.entry_factura.get().strip(),
                descripcion=self.entry_desc.get().strip(),
                unidad=self.entry_unidad.get().strip(),
                cantidad=cantidad,
                precio_sin_iva_bs=precio_sin,
                precio_con_iva_bs=precio_con,
                precio_sin_iva_usd=precio_sin_usd,
                precio_con_iva_usd=precio_con_usd,
                fecha=self.entry_fecha.get().strip(),
                total_sin_iva=total_sin,
                proveedor=self.entry_proveedor.get().strip(),
                proyecto=self.proyecto,
            )
            if self.registro_en_edicion is not None:
                self.registros[self.registro_en_edicion] = registro
            else:
                self.registros.append(registro)

        else:  # Mano de obra
            try:
                monto_usd = float(self.entry_monto_mano_usd.get())
            except ValueError:
                self._mostrar_error("Monto manoobra debe ser numérico en USD.")
                return
            gastos_usd = 0.0
            if self.entry_gastos_extras_usd.get().strip():
                try:
                    gastos_usd = float(self.entry_gastos_extras_usd.get())
                except ValueError:
                    self._mostrar_error("Gastos extras debe ser numérico en USD.")
                    return
            if self.precio_dolar <= 0:
                messagebox.showwarning("Atención", "Primero ingresa el precio del dólar del día.")
                return
            monto_bs = round(monto_usd * self.precio_dolar, 2)
            gastos_bs = round(gastos_usd * self.precio_dolar, 2)
            total_sin = monto_bs
            semana = self.entry_semana.get().strip()
            registro = RegistroGasto(
                item=len(self.registros) + 1,
                tipo="Mano de obra",
                factura="",
                descripcion=self.entry_desc.get().strip(),
                unidad="",
                cantidad=1.0,
                precio_sin_iva_bs=monto_bs,
                precio_con_iva_bs=monto_bs,
                precio_sin_iva_usd=monto_usd,
                precio_con_iva_usd=monto_usd,
                fecha=self.entry_fecha.get().strip(),
                total_sin_iva=total_sin,
                proveedor=self.entry_proveedor.get().strip(),
                proyecto=self.proyecto,
                semana=semana,
                monto_manoobra_usd=monto_usd,
                gastos_extras_usd=gastos_usd,
            )
            if self.registro_en_edicion is not None:
                self.registros[self.registro_en_edicion] = registro
            else:
                self.registros.append(registro)

        self._recalcular_items()
        self._guardar_datos()
        self._limpiar_campos()
        self._refrescar_tabla()
        messagebox.showinfo("Éxito", "Registro guardado correctamente.")

    def _eliminar_gasto(self) -> None:
        seleccion = self.tabla.selection()
        if not seleccion:
            messagebox.showwarning("Atención", "Selecciona uno o más registros para eliminar.")
            return
        if not messagebox.askyesno("Confirmar", "¿Deseas eliminar los registros seleccionados?"):
            return
        try:
            items_a_eliminar = {int(self.tabla.item(s)["values"][0]) for s in seleccion}
        except Exception:
            messagebox.showerror("Error", "No se pudo determinar los registros seleccionados.")
            return
        self.registros = [r for r in self.registros if r.item not in items_a_eliminar]
        self._recalcular_items()
        self._guardar_datos()
        self._limpiar_campos()
        self._refrescar_tabla()

    def _limpiar_lista(self) -> None:
        if not self.registros:
            return
        if messagebox.askyesno("Confirmar", "Esto eliminará todos los registros de este proyecto. ¿Deseas continuar?"):
            self.registros.clear()
            self._guardar_datos()
            self._limpiar_campos()
            self._refrescar_tabla()

    def _iniciar_edicion(self, event: tk.Event[tk.Misc]) -> None:
        seleccionado = self.tabla.focus()
        if not seleccionado:
            return
        valores = self.tabla.item(seleccionado)["values"]
        if not valores:
            return
        item_id = int(valores[0])
        indice = item_id - 1
        if indice < 0 or indice >= len(self.registros):
            return
        registro = self.registros[indice]
        self.registro_en_edicion = indice

        # set mode and populate fields
        self.combo_modo.set(registro.tipo)
        self._aplicar_estado_por_modo()
        try:
            self.entry_factura.delete(0, tk.END)
            self.entry_factura.insert(0, registro.factura)
            self.entry_desc.delete(0, tk.END)
            self.entry_desc.insert(0, registro.descripcion)
            self.entry_unidad.delete(0, tk.END)
            self.entry_unidad.insert(0, registro.unidad)
            self.entry_cantidad.delete(0, tk.END)
            self.entry_cantidad.insert(0, str(registro.cantidad))
            self.entry_precio_sin_iva.delete(0, tk.END)
            self.entry_precio_sin_iva.insert(0, str(registro.precio_sin_iva_bs))
            self.entry_precio_con_iva.delete(0, tk.END)
            self.entry_precio_con_iva.insert(0, str(registro.precio_con_iva_bs))
            self.entry_fecha.delete(0, tk.END)
            self.entry_fecha.insert(0, registro.fecha)
            self.entry_proveedor.delete(0, tk.END)
            self.entry_proveedor.insert(0, registro.proveedor)
            self.entry_semana.delete(0, tk.END)
            self.entry_semana.insert(0, registro.semana)
            self.entry_monto_mano_usd.delete(0, tk.END)
            self.entry_monto_mano_usd.insert(0, str(registro.monto_manoobra_usd))
            self.entry_gastos_extras_usd.delete(0, tk.END)
            self.entry_gastos_extras_usd.insert(0, str(registro.gastos_extras_usd))
        except Exception:
            pass
        self.btn_agregar.config(text="Actualizar gasto")

    def _aplicar_filtros(self) -> None:
        texto_busqueda = self.entry_buscar.get().strip().lower()
        proveedor = self.entry_filtro_proveedor.get().strip().lower()
        fecha_desde = self.entry_fecha_desde.get().strip()
        fecha_hasta = self.entry_fecha_hasta.get().strip()
        monto_min = self.entry_monto_min.get().strip()
        monto_max = self.entry_monto_max.get().strip()

        fecha_inicio_dt = self._parse_fecha(fecha_desde) if fecha_desde else None
        fecha_fin_dt = self._parse_fecha(fecha_hasta) if fecha_hasta else None
        if (fecha_desde and not fecha_inicio_dt) or (fecha_hasta and not fecha_fin_dt):
            self._mostrar_error("Fechas de filtro inválidas. Use DD/MM/YYYY")
            return

        try:
            monto_min_val = float(monto_min) if monto_min else None
            monto_max_val = float(monto_max) if monto_max else None
        except ValueError:
            self._mostrar_error("Montos de filtro inválidos.")
            return

        filtrados: list[RegistroGasto] = []
        for registro in self.registros:
            if texto_busqueda:
                texto = f"{registro.factura} {registro.proveedor} {registro.descripcion}".lower()
                if texto_busqueda not in texto:
                    continue
            if proveedor and proveedor not in registro.proveedor.lower():
                continue
            fecha_dt = self._parse_fecha(registro.fecha)
            if fecha_inicio_dt and (not fecha_dt or fecha_dt < fecha_inicio_dt):
                continue
            if fecha_fin_dt and (not fecha_dt or fecha_dt > fecha_fin_dt):
                continue
            if monto_min_val is not None and registro.total_sin_iva < monto_min_val:
                continue
            if monto_max_val is not None and registro.total_sin_iva > monto_max_val:
                continue
            filtrados.append(registro)

        self._refrescar_tabla(filtrados)

    @staticmethod
    def _parse_fecha(fecha: str) -> Optional[dt.datetime]:
        try:
            return dt.datetime.strptime(fecha, "%d/%m/%Y")
        except Exception:
            return None

    def _limpiar_filtros(self) -> None:
        for entry in [
            self.entry_buscar,
            self.entry_filtro_proveedor,
            self.entry_fecha_desde,
            self.entry_fecha_hasta,
            self.entry_monto_min,
            self.entry_monto_max,
        ]:
            entry.delete(0, tk.END)
        self._refrescar_tabla()

    def _refrescar_tabla(self, registros: Optional[list[RegistroGasto]] = None) -> None:
        registros = registros if registros is not None else self.registros
        for row in self.tabla.get_children():
            self.tabla.delete(row)
        for registro in registros:
            self.tabla.insert("", "end", values=registro.to_row())
        self._actualizar_resumen(registros)

    def _actualizar_resumen(self, registros: list[RegistroGasto]) -> None:
        total_con_iva_bs = sum(r.precio_con_iva_bs * (r.cantidad if r.cantidad else 1) for r in registros)
        total_con_iva_usd = round(total_con_iva_bs / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0
        self.label_total_bs.config(text=f"Total Bs (con IVA): {total_con_iva_bs:,.2f}")
        self.label_total_usd.config(text=f"Total $ (con IVA): {total_con_iva_usd:,.2f}")
        self.label_proyecto.config(text=f"Proyecto: {self.proyecto}")
        self.label_cantidad.config(text=f"Registros: {len(registros)}")

    def _exportar_excel(self) -> None:
        if not self.registros:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return
        try:
            presupuesto_usd = float(self.entry_presupuesto.get())
        except Exception:
            messagebox.showerror("Error", "Debes ingresar un presupuesto válido en dólares.")
            return
        if self.precio_dolar <= 0:
            messagebox.showerror("Error", "Primero ingresa el precio del dólar del día para convertir el presupuesto.")
            return

        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivo Excel", "*.xlsx")],
                                               title="Guardar reporte de gastos",
                                               initialfile=f"reporte_gastos_{self.proyecto}.xlsx")
        if not archivo:
            return

        wb = Workbook()
        ws = wb.active

        titulo = f"REPORTE DE GASTOS - {self.proyecto} - {dt.datetime.now().strftime('%d/%m/%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
        celda_titulo = ws.cell(row=1, column=1, value=titulo)
        celda_titulo.font = Font(size=16, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")

        header_row = 3
        for col_idx, encabezado in enumerate(COLUMNAS, start=1):
            celda = ws.cell(row=header_row, column=col_idx, value=encabezado)
            celda.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            celda.font = Font(color="FFFFFF", bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, registro in enumerate(self.registros, start=header_row + 1):
            for col_idx, valor in enumerate(registro.to_row(), start=1):
                celda = ws.cell(row=row_idx, column=col_idx, value=valor)
                celda.alignment = Alignment(horizontal="center", vertical="center")

        total_sin_iva_bs = sum(r.precio_sin_iva_bs * (r.cantidad if r.cantidad else 1) for r in self.registros)
        total_con_iva_bs = sum(r.precio_con_iva_bs * (r.cantidad if r.cantidad else 1) for r in self.registros)
        total_con_iva_usd = round(total_con_iva_bs / self.precio_dolar, 2) if self.precio_dolar > 0 else 0.0

        total_fila = header_row + len(self.registros) + 2

        ws.cell(row=total_fila, column=1, value="PRESUPUESTO (USD):")
        ws.cell(row=total_fila, column=2, value=presupuesto_usd)
        ws.cell(row=total_fila, column=3, value=f"Equiv. Bs (@{self.precio_dolar}):")
        ws.cell(row=total_fila, column=4, value=round(presupuesto_usd * self.precio_dolar, 2))

        ws.cell(row=total_fila + 1, column=1, value="TOTAL SIN IVA (Bs):")
        ws.cell(row=total_fila + 1, column=2, value=round(total_sin_iva_bs, 2))

        ws.cell(row=total_fila + 2, column=1, value="TOTAL CON IVA (Bs):")
        ws.cell(row=total_fila + 2, column=2, value=round(total_con_iva_bs, 2))

        utilidad_usd = round(presupuesto_usd - total_con_iva_usd, 2)
        porcentaje = (utilidad_usd / presupuesto_usd * 100) if presupuesto_usd > 0 else 0

        ws.cell(row=total_fila + 3, column=1, value="UTILIDAD (USD):")
        ws.cell(row=total_fila + 3, column=2, value=utilidad_usd)

        ws.cell(row=total_fila + 4, column=1, value="% UTILIDAD:")
        ws.cell(row=total_fila + 4, column=2, value=round(porcentaje, 2))

        for i in range(total_fila, total_fila + 5):
            for j in range(1, 5):
                celda = ws.cell(row=i, column=j)
                celda.fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center", vertical="center")

        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws.iter_rows(min_row=header_row, max_row=total_fila + 4, max_col=len(COLUMNAS)):
            for celda in row:
                celda.border = thin

        for col_idx in range(1, len(COLUMNAS) + 1):
            max_len = 0
            for row in range(1, total_fila + 5):
                texto = str(ws.cell(row=row, column=col_idx).value or "")
                if len(texto) > max_len:
                    max_len = len(texto)
            ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = max_len + 4

        wb.save(archivo)
        wb.close()
        messagebox.showinfo("Éxito", "Reporte generado con totales (sin/ con IVA), presupuesto en USD y utilidad.")

    def _seleccionar_proyecto_modal(self) -> None:
        proyectos = []
        if hasattr(self, "_df_global") and not self._df_global.empty and "Proyecto" in self._df_global.columns:
            proyectos = sorted(self._df_global["Proyecto"].dropna().unique().tolist())

        modal = tk.Toplevel(self.root)
        modal.title("Seleccionar o crear proyecto")
        modal.geometry("420x160")
        modal.transient(self.root)
        modal.grab_set()

        ttk.Label(modal, text="Selecciona un proyecto existente:").pack(anchor="w", padx=10, pady=(10, 2))
        combo = ttk.Combobox(modal, values=proyectos, state="readonly" if proyectos else "normal", width=40)
        combo.pack(padx=10, pady=2)
        if proyectos:
            combo.set(proyectos[0])

        ttk.Label(modal, text="O crea uno nuevo:").pack(anchor="w", padx=10, pady=(8, 2))
        entry_new = ttk.Entry(modal, width=44)
        entry_new.pack(padx=10, pady=2)

        def aceptar():
            elegido = entry_new.get().strip() or combo.get().strip()
            if not elegido:
                messagebox.showerror("Error", "Debes ingresar o seleccionar un nombre de proyecto.")
                return
            self.proyecto = elegido
            if not hasattr(self, "_df_global") or self._df_global is None:
                self._df_global = pd.DataFrame()
            modal.destroy()
            try:
                self._cargar_datos_proyecto()
            except Exception:
                pass
            try:
                self.label_proyecto.config(text=f"Proyecto: {self.proyecto}")
                self.root.title(f"Control de Gastos - Terra Caliza - {self.proyecto}")
            except Exception:
                pass

        btn = ttk.Button(modal, text="Aceptar", command=aceptar)
        btn.pack(pady=8)

        self.root.wait_window(modal)

    def ejecutar(self) -> None:
        self.root.mainloop()


if __name__ == "__main__":
    app = GestorGastos()
    app.ejecutar()
